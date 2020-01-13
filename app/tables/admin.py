from datetime import date
from django.contrib import admin
from django.shortcuts import render, redirect
from django.urls import path
import openpyxl
from rangefilter.filter import DateRangeFilter
from django.contrib import messages
from .forms import XLSXImportForm
from .models import *

# Register your models here.


class NutrientInline(admin.StackedInline):
    model = Nutrient


class TableAdmin(admin.ModelAdmin):
    model = Table
    search_fields = ['date', 'dietary_composition']
    change_list_template = 'admin/tables/Table/change_list.html'
    inlines = [NutrientInline]
    list_filter = (
        ('date', DateRangeFilter),
    )

    # XLSX 파일을 Import 하여 실행할 기능을 정의한다.
    # Import XLSX
    def import_xlsx(self, request):
        if request.method == "POST":
            file = request.FILES.get('xlsx_file')
            initial_year = int(request.POST.get('date_input_year'))
            initial_month = int(request.POST.get('date_input_month'))

            xlsx_file = openpyxl.load_workbook(file)
            nut_sheet = xlsx_file['Nutrition_Sheet']
            des_sheet = xlsx_file['Description_Sheet']
            entry_list = ['전체', '아침', '점심', '저녁', '간식']

            all_meal_list = list(nut_sheet.iter_rows())
            all_meal_des_list = list(des_sheet.iter_rows())

            # find non blank last row of the nut_sheet
            last_nut_row = 0
            for cell in list(nut_sheet.columns)[0]:
                if cell.value is None:
                    last_nut_row = cell.row
                    break
            if last_nut_row:
                num_rows = last_nut_row
            else:
                num_rows = nut_sheet.max_row

            bookmark = 1
            for i in range((num_rows - 1)//5):
                for mark, entry in enumerate(entry_list, start=bookmark):
                    # Table Create
                    if entry is not '전체':
                        try:
                            meal_content = all_meal_des_list[mark - (i + 1)][3].value
                            if ',' in meal_content:
                                composition = meal_content.split(',')
                            else:
                                composition = [meal_content]
                            table, nop = Table.objects.get_or_create(
                                dietary_composition=composition,
                                ingredients=all_meal_des_list[mark - (i + 1)][4].value,
                                recipe=all_meal_des_list[mark - (i + 1)][5].value,
                                tips=all_meal_des_list[mark - (i + 1)][6].value,
                                date=date(
                                    initial_year,
                                    initial_month,
                                    i+1
                                ),
                                time=entry
                            )
                            # Table Nutrients Create
                            nut = table.nutrient
                            nut.calorie = all_meal_list[mark][3].value
                            nut.carbs = all_meal_list[mark][4].value
                            nut.fiber = all_meal_list[mark][5].value
                            nut.V_protein = all_meal_list[mark][6].value
                            nut.A_protein = all_meal_list[mark][7].value
                            nut.V_fat = all_meal_list[mark][8].value
                            nut.A_fat = all_meal_list[mark][9].value
                            nut.cholesterol = all_meal_list[mark][10].value
                            nut.salt = all_meal_list[mark][11].value
                            nut.potassium = all_meal_list[mark][12].value
                            nut.phosphorus = all_meal_list[mark][13].value
                            nut.V_calcium = all_meal_list[mark][14].value
                            nut.A_calcium = all_meal_list[mark][15].value
                            nut.V_iron = all_meal_list[mark][16].value
                            nut.A_iron = all_meal_list[mark][17].value
                            nut.save()

                        except ValueError:
                            messages.add_message(
                                request,
                                messages.WARNING,
                                '식단의 수가 생성하려는 달의 날짜 수와 일치하지 않습니다.'
                            )
                            return redirect(".")
                        except AttributeError:
                            messages.add_message(
                                request,
                                messages.WARNING,
                                '엑셀 파일의 내용 중 형식이 맞지 않는 컨텐츠가 존재합니다.'
                            )
                            return redirect('.')
                    bookmark += 1
            return redirect("..")
        # XLSX file 제출용 HTML 페이지로 연결
        # 여기서 XLSX 파일을 업로드하면
        form = XLSXImportForm()
        return render(
            request, "admin/xlsx_form.html", {"form": form}
        )

    def get_urls(self):
        urls = super(TableAdmin, self).get_urls()
        new_urls = [
            path('import-xlsx/', self.import_xlsx),
        ] + urls
        return new_urls


class TableLogAdmin(admin.ModelAdmin):
    model = TableLog
    change_list_template = 'admin/tables/TableLog/change_list.html'
    search_fields = [
        'user__username', 'user__name', 'date'
    ]
    list_display = ['user', 'date', 'time']
    list_filter = (
        ('date', DateRangeFilter),
    )


admin.site.register(Table, TableAdmin)
admin.site.register(TableLog, TableLogAdmin)